"""
Document parser: PDF (pdfplumber), Word (python-docx), Excel/CSV (openpyxl).
Returns page-level or sheet-level text for chunking. No raw text written to disk.
"""
from __future__ import annotations

import csv
import io
import logging
from typing import TypedDict

import pdfplumber
from docx import Document as DocxDocument
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


class PageText(TypedDict):
    page_number: int
    text: str


def parse_pdf(data: bytes) -> list[PageText]:
    """Extract text per page. Uses pdfplumber (do not install pdfminer.six separately)."""
    out: list[PageText] = []
    with io.BytesIO(data) as buf:
        with pdfplumber.open(buf) as pdf:
            for i, page in enumerate(pdf.pages, start=1):
                text = page.extract_text()
                out.append({"page_number": i, "text": text or ""})
    return out


def parse_docx(data: bytes) -> list[PageText]:
    """Extract paragraphs. Docx has no real page breaks; treat as single logical page 1."""
    with io.BytesIO(data) as buf:
        doc = DocxDocument(buf)
        parts = [p.text for p in doc.paragraphs if p.text.strip()]
        text = "\n\n".join(parts)
    return [{"page_number": 1, "text": text}]


def parse_xlsx(data: bytes) -> list[PageText]:
    """Extract text from all sheets. Each sheet = logical page."""
    out: list[PageText] = []
    with io.BytesIO(data) as buf:
        wb = load_workbook(buf, read_only=True, data_only=True)
        for sheet_idx, sheet in enumerate(wb.worksheets, start=1):
            rows: list[str] = []
            for row in sheet.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                rows.append("\t".join(cells))
            text = "\n".join(rows)
            out.append({"page_number": sheet_idx, "text": text})
        wb.close()
    return out


def parse_csv(data: bytes) -> list[PageText]:
    """Single logical page (page_number=1)."""
    text = data.decode("utf-8", errors="replace")
    with io.StringIO(text) as buf:
        reader = csv.reader(buf)
        rows = ["\t".join(row) for row in reader]
    return [{"page_number": 1, "text": "\n".join(rows)}]


def parse_document(data: bytes, file_type: str) -> list[PageText]:
    """Dispatch by file_type (pdf, docx, xlsx, csv). Raises ValueError for unsupported type."""
    ft = (file_type or "").lower().strip()
    if ft == "pdf":
        return parse_pdf(data)
    if ft == "docx":
        return parse_docx(data)
    if ft == "xlsx":
        return parse_xlsx(data)
    if ft == "csv":
        return parse_csv(data)
    raise ValueError(f"Unsupported file_type: {file_type}")
